from openpyxl import Workbook

exc=Workbook()
exl=exc.active

exl.cell(row=1,column=1,value='Derman_id')
exl.cell(row=1,column=2,value='Derman_Adi')
exl.cell(row=1,column=3,value='Derman_Qiymeti')
exl.cell(row=1,column=4,value='Derman_Sayi')

def insert_row(derman_adi,derman_qiymeti,derman_sayi):
    exl.append((exl.max_row,derman_adi,derman_qiymeti,derman_sayi))
    return list(exl.values)
insert_row('Aspirin',25,50)
insert_row('Spazmalgon',5,100)
exc.save("Pharmacy.xlsx")



